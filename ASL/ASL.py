import os
from sympy import Line,Point
from sympy import abc
from matplotlib import figure
import tkinter
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg ,NavigationToolbar2Tk
import numpy
import random
from PIL import ImageTk
from PIL import Image
import openpyxl
import sqlite3
import pandas
# select atom name & character
def check_atom(name):
        wb = openpyxl.load_workbook("Asl\\element.xlsx")
        ws = wb.active
        for i in range(3, 121):
            if ws.cell(row=i, column=2).value == name:
                numper=ws.cell(row=i, column=1).value
                return print("this element is exist", name,numper)
            if i == 120 and not (ws.cell(row=i, column=2).value == name):
                print("ERROR:[ NO NAME OF ELEMENT LIKE THIS '{}']".format(name))
                return
#-----------------------------------------------------
##controls
def mk_st(name_of_structure):
    conn=sqlite3.connect("ASL\st.db")
    exe=conn.cursor()
    name=str(name_of_structure)+"_"+"Atoms"
    exe.execute("CREATE TABLE {}(atom_name TEXT,numper_of_atom TEXT,no_electron_shell TEXT)".format(name))
    conn.commit()
    name2=str(name_of_structure)+"_"+"angles"
    exe.execute("CREATE TABLE {}(atom_1_name TEXT,numper_of_atom1 TEXT,atom_2_name TEXT,numper_of_atom2 TEXT,angle_between TEXT)".format(name2))
    conn.commit()
    name3=str(name_of_structure)+"_"+"bonds"
    exe.execute("CREATE TABLE {}(atom_1_name TEXT,numper_of_atom1 TEXT,atom_2_name TEXT,numper_of_atom2 TEXT,bond_direction TEXT,type_of_bond TEXT,length TEXT)".format(name3))
    conn.commit()
    conn.close()
    #---------------------
    conn2 = sqlite3.connect("ASL\saved_names.db")
    exe2 = conn2.cursor()
    name2 = str(name_of_structure)
    exe2.execute("""INSERT INTO saved_names VALUES("{}")""".format(name2))
    conn2.commit()
    return print("your structure go in database")
class add_element():
    def add_atom(self,structure_name,atom_name,numper_of_atom,no_electoron_in_shell):
        conn = sqlite3.connect("ASL\st.db")
        exe = conn.cursor()
        name_of_db=str(structure_name)+"_"+"Atoms"
        exe.execute("""INSERT INTO {}(atom_name,numper_of_atom,no_electron_shell) VALUES ("{}","{}","{}") """.format(name_of_db,atom_name,numper_of_atom,no_electoron_in_shell))
        conn.commit()
        conn.close()
        return print("Atom" ,"called","{}".format(atom_name),"add","in","{}".format(structure_name))
    def add_angles(self,structure_name,atom_1_name,numper_of_atom_1,atom_2_name,numper_of_atom_2,angle_between):
        conn = sqlite3.connect("ASL\st.db")
        exe = conn.cursor()
        name_of_db = str(structure_name) + "_" + "angles"
        #angle must be 90 or 80 or 70 ... etc
        exe.execute("""INSERT INTO {}(atom_1_name,numper_of_atom1,atom_2_name,numper_of_atom2,angle_between) VALUES ("{}","{}","{}","{}","{}") """.format(name_of_db,atom_1_name,numper_of_atom_1,atom_2_name,numper_of_atom_2,angle_between))
        conn.commit()
        conn.close()
        return print("Atom", "called", "{}".format(atom_1_name), "make", "angle","with", "atom called","{}".format(atom_2_name),"in","structure","called","{}".format(structure_name))
    def add_bond(self,structure_name,atom_1_name,numper_of_atom_1,atom_2_name,numper_of_atom_2,bond_type,position_of_bond,length):
        conn = sqlite3.connect("ASL\st.db")
        exe = conn.cursor()
        #bondtype{single:1 , double:2 ,triple:3,...etc}
        #bond position{on the line : 0 , above line : 1 , below the line -1}
        name_of_db = str(structure_name) + "_" + "bonds"
        exe.execute("""INSERT INTO {}(atom_1_name,numper_of_atom1,atom_2_name,numper_of_atom2,bond_direction,type_of_bond,length) VALUES ("{}","{}","{}","{}","{}","{}",{}) """.format(name_of_db, atom_1_name, numper_of_atom_1, atom_2_name, numper_of_atom_2,bond_type,position_of_bond,length))
        conn.commit()
        conn.close()
        return print("Atom", "called", "{}".format(atom_1_name), "make", "bond", "with", "atom called","{}".format(atom_2_name), "in", "structure", "called", "{}".format(structure_name))
class del_element():
    def del_atom(self,st_name,atom_name,numper_of_atom):
        conn = sqlite3.connect("ASL\st.db")
        exe = conn.cursor()
        name = str(st_name) + "_" + "Atoms"
        exe.execute("DELETE  FROM {} WHERE atom_name='{}' AND numper_of_atom='{}' ".format(name,atom_name,numper_of_atom))
        conn.commit()
        conn.close()
        return print("Atom", "called", "{}".format(atom_name),"Numper is {}".format(numper_of_atom), "deleted", "from", "{}".format(st_name))
    def del_angle(self,st_name,atom_1_name,numper_of_atom1,atom_2_name,numper_of_atom2):
        conn = sqlite3.connect("ASL\st.db")
        exe = conn.cursor()
        name = str(st_name) + "_" + "angles"
        exe.execute("DELETE  FROM {} WHERE atom_1_name='{}' AND numper_of_atom1='{}' AND atom_2_name='{}' AND numper_of_atom2='{}' ".format(name, atom_1_name, numper_of_atom1,atom_2_name,numper_of_atom2))
        conn.commit()
        conn.close()
        return print("delete angle between ","Atom", "called", "{}".format(atom_1_name),"Numper is {}".format(numper_of_atom1), "and atom called {} this atom numper {}".format(atom_2_name,numper_of_atom2), "from", "{}".format(st_name))
    def del_bond(self,st_name,atom_1_name,numper_of_atom1,atom_2_name,numper_of_atom2):
        conn = sqlite3.connect("ASL\st.db")
        exe = conn.cursor()
        name = str(st_name) + "_" + "bonds"
        exe.execute("DELETE  FROM {} WHERE atom_1_name='{}' AND numper_of_atom1='{}' AND atom_2_name='{}' AND numper_of_atom2='{}' ".format(name, atom_1_name, numper_of_atom1, atom_2_name, numper_of_atom2))
        conn.commit()
        conn.close()
        ###

        return print(" delete bond between ","Atom", "called", "{}".format(atom_1_name),"Numper is {}".format(numper_of_atom1), "and atom called {} this atom numper {}".format(atom_2_name,numper_of_atom2), "from", "{}".format(st_name))
    def drop_st(self,name_of_st):
        conn = sqlite3.connect("ASL\st.db")
        exe = conn.cursor()
        name = str(name_of_st) + "_" + "Atoms"
        exe.execute("DROP TABLE {}".format(name))
        conn.commit()
        name2 = str(name_of_st) + "_" + "angles"
        exe.execute( "DROP TABLE {}".format(name2))
        conn.commit()
        name3 = str(name_of_st) + "_" + "bonds"
        exe.execute("DROP TABLE {}".format(name3))
        conn.commit()
        conn.close()
        #
        conn = sqlite3.connect("ASL\saved_names.db")
        exe = conn.cursor()
        name = str(name_of_st)
        exe.execute("DELETE  FROM {} WHERE name='{}'  ".format("saved_names",name))
        conn.commit()
        conn.close()
        ###
        return print("{}-{}-{} droped".format(name,name2,name3))
##tools to calc
def type_of_atoms_table(st_name):
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    name = str(st_name) + "_" + "Atoms"

    exe.execute("""SELECT atom_name FROM {}""".format(name))
    n=0
    add_list=[]
    list1_name=[]
    list2_count=[]
    percentage=[]
    for res in exe.fetchall():
        n=n+1
        add_list.append(res[0])
    #open exel
    exel=openpyxl.load_workbook("./Asl/element.xlsx")
    wb=exel.active
    for count in range(1,wb.max_row):
        if add_list.count("{}".format(str(wb.cell(row=count , column=2).value)))>1:
            list1_name.append(wb.cell(row=count , column=2).value)
            list2_count.append(add_list.count("{}".format(str(wb.cell(row=count , column=2).value))))
        if add_list.count("{}".format(str(wb.cell(row=count, column=2).value))) == 1:
            list1_name.append(wb.cell(row=count, column=2).value)
            list2_count.append(add_list.count("{}".format(str(wb.cell(row=count, column=2).value))))
    sum=[list1_name,list2_count]
    pand=pandas.DataFrame(sum,index=["atom_name","atom_count"])
    return list2_count
def type_of_atoms_table_show(st_name):
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    name = str(st_name) + "_" + "Atoms"

    exe.execute("""SELECT atom_name FROM {}""".format(name))
    n=0
    add_list=[]
    list1_name=[]
    list2_count=[]
    percentage=[]
    for res in exe.fetchall():
        n=n+1
        add_list.append(res[0])
    #open exel
    exel=openpyxl.load_workbook("./Asl/element.xlsx")
    wb=exel.active
    for count in range(1,wb.max_row):
        if add_list.count("{}".format(str(wb.cell(row=count , column=2).value)))>1:
            list1_name.append(wb.cell(row=count , column=2).value)
            list2_count.append(add_list.count("{}".format(str(wb.cell(row=count , column=2).value))))
        if add_list.count("{}".format(str(wb.cell(row=count, column=2).value))) == 1:
            list1_name.append(wb.cell(row=count, column=2).value)
            list2_count.append(add_list.count("{}".format(str(wb.cell(row=count, column=2).value))))
    sum=[list1_name,list2_count]
    pand=pandas.DataFrame(sum,index=["atom_name","atom_count"])
    print(pand)
    return 0

def bond_DIRECTION_calc(structure_name):
    tup=[]
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    # bondtype{single:1 , double:2 ,triple:3,...etc}
    # bond position{on the line : 0 , above line : 1 , below the line -1}
    name_of_db = str(structure_name) + "_" + "bonds"
    exe.execute("""SELECT bond_direction FROM {} """.format(name_of_db))
    for new_rec in exe.fetchall():
        tup.append(new_rec[0])
    final_tuple = check_tuple(tup)
    return final_tuple
def bond_TYPE_calc(structure_name):
    tup = []
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    # bondtype{single:1 , double:2 ,triple:3,...etc}
    # bond position{on the line : 0 , above line : 1 , below the line -1}
    name_of_db = str(structure_name) + "_" + "bonds"
    exe.execute("""SELECT type_of_bond FROM {} """.format(name_of_db))
    for new_rec in exe.fetchall():
        tup.append(new_rec[0])
    final_tuple = check_tuple(tup)
    return final_tuple
def bond_length_calc(structure_name):
    tup = []
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    # bondtype{single:1 , double:2 ,triple:3,...etc}
    # bond position{on the line : 0 , above line : 1 , below the line -1}
    name_of_db = str(structure_name) + "_" + "bonds"
    exe.execute("""SELECT length FROM {} """.format(name_of_db))
    for new_rec in exe.fetchall():
        tup.append(new_rec[0])
    conn.close()
    final_tuple = check_tuple(tup)
    return final_tuple
def angles_calc(structure_name):
    tup = []
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    name_of_db = str(structure_name) + "_" + "angles"
    exe.execute("""SELECT angle_between FROM {} """.format(name_of_db))
    for new_rec in exe.fetchall():
        tup.append(new_rec[0])
    final_tuple=check_tuple(tup)
    return final_tuple
def total_electron_shell_calc(structure_name):
    tup = []
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    name_of_db = str(structure_name) + "_" + "Atoms"
    exe.execute("""SELECT no_electron_shell FROM {} """.format(name_of_db))
    for new_rec in exe.fetchall():
        tup.append(new_rec[0])
    final_tuple = check_tuple(tup)
    val = 0
    v=0
    calculate="...."
    for u in final_tuple[0:]:
        v = v + 1
        val += float(u)

    return val
def check_tuple(tupe_values):
    final_tuple=[]
    for tt in range(0,len(tupe_values)):
        if tupe_values[tt]=="none" or tupe_values[tt]=="None" or tupe_values[tt]==None:
            mm=tupe_values[tt]="0"
            final_tuple.append(mm)
        final_tuple.append(tupe_values[tt])
    return final_tuple
def vasperAngle_img_show():
    tk=tkinter.Tk()
    tk.title("vasperAngles")
    open1=Image.open("ASL//imageedit_57_3142624754.jpg")
    show1 = ImageTk.PhotoImage(open1)
    open2 = Image.open("ASL//imageedit_132_6645880028.jpg")
    show2 = ImageTk.PhotoImage(open2)
    open3 = Image.open("ASL//vsepr05.png")
    show3 = ImageTk.PhotoImage(open3)
    tkinter.Label(image=show1).place(x=0,y=0)
    tkinter.Label(image=show2).place(x=0,y=350)
    tkinter.Label(image=show3).place(x=1150,y=350)
    tk.mainloop()
def total_numper_of_atom(st_name):
    count=[]
    a_numper=sqlite3.connect("ASL\st.db")
    exee=a_numper.cursor()
    database_name =str(st_name) + "_" + "Atoms"
    exee.execute(""" SELECT atom_name FROM {}   """.format(database_name))
    for n in exee.fetchall():
        count.append(n[0:])
    return len(count)
def tuple_precessing(tup1,tup2):
    if len(tup1)>len(tup2):
        sum=len(tup1)-len(tup2)
        for rep in range(0,sum):
            tup2.append("E")
        return tup2
    if len(tup2) > len(tup1):
        sum = len(tup2) - len(tup1)
        for rep in range(0, sum):
            tup1.append("E")
        return tup1
            # ---
def get_perc_result(no1,no2):
    if float(no1)>float(no2):
        sum=(float(no2)/float(no1))*100
        return float(sum)
    if float(no1) < float(no2):
        sum = (float(no1) / float(no2)) * 100
        return float(sum)
    if float(no1) == float(no2):
        sum="100"
        return float(sum)
def process_E_value(value):
    if value=="E":
        return int(0)
    else:
        return float(value)
def solve_tuple_calc(tup_bond1,tup_bond2):
    final_numper_tup=[]
    #precess
    if len(tup_bond1)>len(tup_bond2):
        modfy_tup_2=tuple_precessing(tup_bond1,tup_bond2)

        for count_result in range(0,len(modfy_tup_2)):

            ev1=process_E_value(str(tup_bond1[count_result]))

            ev2 = process_E_value(str(tup_bond2[count_result]))

            final_numper_tup.append(get_perc_result(ev1,ev2))
    if len(tup_bond2) > len(tup_bond1):

        modfy_tup_1 = tuple_precessing(tup_bond1, tup_bond2)
        for count_result in range(0, len(modfy_tup_1)):


            ev1 = process_E_value(str(tup_bond1[count_result]))

            ev2 = process_E_value(str(tup_bond2[count_result]))

            final_numper_tup.append(get_perc_result(ev1, ev2))
    numper1=len(final_numper_tup)
    final_numper=int(numper1)*100
    #sum thw tupe
    vol=0
    for count in final_numper_tup[0:]:
        vol=count+vol
    rr=get_perc_result(vol,final_numper)
    return rr
#bond direction and type calculations
def bond_type_di_static(st_name1,st_name2):
    #1
    bond1_di=[]
    bond1_type = []
    bond2_di = []
    bond2_type = []
    result_=[]
    name_of_db = str(st_name1) + "_" + "bonds"
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    exe.execute("""SELECT bond_direction FROM {} """.format(name_of_db))
    for add_in_tu_bond1_di in exe.fetchall():
        bb=bond1_di.append(add_in_tu_bond1_di[0])
    exe.execute("""SELECT type_of_bond FROM {} """.format(name_of_db))
    for add_in_tu_bond1_type in exe.fetchall():
        bb = bond1_type.append(add_in_tu_bond1_type[0])
    ##modification

    #---------------------------
    #2
    name_of_db = str(st_name2) + "_" + "bonds"
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    exe.execute("""SELECT bond_direction FROM {} """.format(name_of_db))
    for add_in_tu_bond2_di in exe.fetchall():
        bb = bond2_di.append(add_in_tu_bond2_di[0])
    exe.execute("""SELECT type_of_bond FROM {} """.format(name_of_db))
    for add_in_tu_bond2_type in exe.fetchall():
        bb = bond2_type.append(add_in_tu_bond2_type[0])
    # result of di
    accuri=0
    for rs in bond1_di[0:]:
        for rs2 in bond2_di[0:]:
            if rs[0]==rs2[0]:
                accuri=accuri+1

        #percentage
        persentage_di=get_perc_result(len(bond2_di),accuri)
        string1=str(persentage_di)
        result_.append(string1)
    accuri_ty=0
    #result of type
    mm=[]
    for rs in bond1_type[0:]:
        for rs2 in bond2_type[0:]:
            if rs[0] == rs2[0]:
                accuri_ty = accuri_ty + 1

        # percentage
        persentage_type = get_perc_result(len(bond2_type), accuri_ty)
        string2=str(persentage_type)
        result_.append(string2)
        #CALCULATION SUM


    return result_
def RS_decide(val1,val2):
    if str(val1)==str(val2):
        res="similar"
        return res
    else:
        res="not similar"
        return res

#---------- collect calc-----------------#
def prepare_to_identify(name_of_st,type_RS_of_structure,total_not_paired_electron,total_paired_electron):
    tupe=(e_numper(name_of_st),mwt(name_of_st),bond_DIRECTION_calc(name_of_st),bond_TYPE_calc(name_of_st),bond_length_calc(name_of_st),angles_calc(name_of_st),total_electron_shell_calc(name_of_st),total_Density_calc(name_of_st),type_RS_of_structure,total_not_paired_electron,total_paired_electron,name_of_st)
    return tupe
#pattern
def find_pattern_numper(no1,no2,rate):
    patt_res=[]
    if rate == 0:
        rate=.000001
    frac=[]
    #we must delete all (-|+) from pattern completly
    pat=[no1,no2]
    t1_result=[]
    x=1
    y=1
    int(x)
    int(y)
    #--how discover this pattern
    ##positive
    #small number
    if float(no1) == float(no2):
        tupe = ["1"]
        print("final", tupe[0])
        patt_res.append(tupe)
        return 1.0



    # large number
    while  not (int((((x / y) * pat[0])/ pat[1])*100)==int(100)) :

        while (int(pat[0]*(x/y))>int(pat[1])):
            rr1=pat[0]*(x/y)
            x=x-rate

            t1_result.append(rr1)
            frac.append(x/y)
            tupe = [x]
            print(tupe)
            print(int((((x / y) * pat[0])/ pat[1])*100))
            if int((((x / y) * pat[0]) / pat[1]) * 100) == int(100):
                patt_res.append(tupe)
                print("final", tupe)
                break



            while int((pat[0] * (x / y)) < int(pat[1])):
                rr1 = pat[0] * (x / y)
                x = x + rate

                frac.append(x / y)
                tupe = [x]
                print(tupe)
                print(int((((x / y) * pat[0])/ pat[1])*100))

                if int((((x / y) * pat[0])/ pat[1])*100)==int(100):
                    patt_res.append(tupe)
                    print("final", tupe)
                    break




        while (int(pat[0] * (x / y)) < int(pat[1])):
            rr1 = pat[0] * (x / y)
            x = x + rate
            frac.append(x/y)
            tupe = [x]
            print(int((((x / y) * pat[0])/ pat[1])*100))
            #print(tupe)
            if int((((x / y) * pat[0]) / pat[1]) * 100) == int(100):
                patt_res.append(tupe)
                print("final", tupe)
                break

                # small ex


            while (int(pat[0] * (x / y)) > int(pat[1])):
                rr1 = pat[0] * (x / y)
                x = x - rate
                t1_result.append(rr1)
                frac.append(x / y)
                tupe1 = [x]
                print(tupe1)
                print(int((((x / y) * pat[0])/ pat[1])*100))
                if int((((x / y) * pat[0])/ pat[1])*100)==int(100):
                    patt_res.append(tupe1)
                    print("final",tupe1)
                    break

                    # small ex

    if float(no1)==float(no2):
        tupe=[1]
        patt_res.append(tupe)
    print("----------------------------------------------------------------")
    return patt_res[0]

def find_pattern_error(no1,no2,rate):
    patt_res = []
    if rate == 0:
        rate = .000001
    frac = []
    # we must delete all (-|+) from pattern completly
    pat = [no1, no2]
    t1_result = []
    x = 1
    y = 1
    int(x)
    int(y)


    # --how discover this pattern
    ##positive
    while not (int((((x / y) * pat[0]) / pat[1]) * 100) == int(100)):

        while (int(pat[0] * (x / y)) > int(pat[1])):
            rr1 = pat[0] * (x / y)
            x = x - rate

            t1_result.append(rr1)
            frac.append(x / y)
            tupe = [(((x/y)*pat[0])/pat[1])*100]
            print(tupe)
            # print(tupe)
            if int((((x / y) * pat[0]) / pat[1]) * 100) == int(100):
                patt_res.append(tupe)
                print("final err8", tupe)
                break


            while int((pat[0] * (x / y)) < int(pat[1])):
                rr1 = pat[0] * (x / y)
                x = x + rate

                frac.append(x / y)
                tupe = [(((x/y)*pat[0])/pat[1])*100]
                print(tupe)

                if int((((x / y) * pat[0]) / pat[1]) * 100) == int(100):
                    patt_res.append(tupe)
                    print("final err6", tupe)
                    break


        while (int(pat[0] * (x / y)) < int(pat[1])):
            rr1 = pat[0] * (x / y)
            x = x + rate
            frac.append(x / y)
            tupe = [(((x/y)*pat[0])/pat[1])*100]
            print(tupe)
            if int((((x / y) * pat[0]) / pat[1]) * 100) == int(100):
                patt_res.append(tupe)
                print("final err4", tupe)
                break

                # small ex


            while (int(pat[0] * (x / y)) > int(pat[1])):
                rr1 = pat[0] * (x / y)
                x = x - rate
                t1_result.append(rr1)
                frac.append(x / y)
                tupe1 = [(((x/y)*pat[0])/pat[1])*100]
                print(tupe1)
                if int((((x / y) * pat[0]) / pat[1]) * 100) == int(100):
                    patt_res.append(tupe1)
                    print("final err2",tupe1)
                    break

                    # small ex

            print("----------------------------------------------------------------")
    if float(no1)==float(no2):
        tupe=[float(100)]
        patt_res.append(tupe)
        return 100.0
    return patt_res[0]

#---------------similarity_____________#
def rational_functionality(no_tu1,no_tu2):
    #e-numper
    func1=get_perc_result(no_tu1[0],no_tu2[0])
    func2=get_perc_result(no_tu1[1], no_tu2[1])
    #angle calc
    func3=solve_tuple_calc(no_tu1[4], no_tu2[4])
    func4=solve_tuple_calc(no_tu1[5], no_tu2[5])
    func5 = get_perc_result(no_tu1[6], no_tu2[6])
    func6 = get_perc_result(no_tu1[7], no_tu2[7])
    func7 = get_perc_result(no_tu1[9], no_tu2[9])
    func8 = get_perc_result(no_tu1[10], no_tu2[10])
    func9=bond_type_di_static(no_tu1[11],no_tu2[11])
    func10=RS_decide(no_tu1[8], no_tu2[8])
    func11=get_perc_result(total_proton_calc(no_tu1[11]),total_proton_calc(no_tu2[11]))
    #bond  type and direction
    print("--------------------SIMILARITY REPORT-----------START----------------")
    print("e-numper","->",int(float(func1)),"%")
    print("mwt", "->",int(float(func2)),"%")
    print("type & direction", "->", int(float(func9[0])),"%","--",int(float(func9[1])),"%")
    print("length", "->", int(float(func3)),"%")
    print("angle", "->", int(float(func4)),"%")
    print("total e shell calc", "->", int(float(func5)),"%")
    print("total Den. shell", "->", int(float(func6)),"%")
    print("total not paired E-", "->", int(float(func7)),"%")
    print("total paired E-", "->", int(float(func8)),"%")
    print("R&S", "->", func10)
    print("total protons", "->", func11)
    print("-----------------------------------------------")
    #similarity
    similarity_tp=[]
    real_similarity=[]
    if func10=="similar":
        converted_value=100
    else:
        converted_value = 0

    #ctotal calculation
    similarity_tp.append(int(float(func1)))
    similarity_tp.append(int(float(func2)))
    similarity_tp.append(int(float(func3)))
    similarity_tp.append(int(float(func9[0])))
    similarity_tp.append(int(float(func9[1])))
    similarity_tp.append(int(float(func4)))
    similarity_tp.append(int(float(func5)))
    similarity_tp.append(int(float(func6)))
    similarity_tp.append(int(float(func7)))
    similarity_tp.append(int(float(func8)))
    similarity_tp.append(int(float(converted_value)))
    similarity_tp.append(int(float(func11)))
    #add atom kind
    v=0
    for v_v in similarity_tp[0:]:
        v=int(v)+int(v_v)
    numper=len(similarity_tp)*100
    sim_res=get_perc_result(v,numper)
    print("**","more than total paired and angles similarity percentage is :- ",sim_res,"%")
    #which is active
    real_similarity.append(int(float(func4)))
    real_similarity.append(int(float(func8)))
    # calculating type numper atoms
    if len(type_of_atoms_table(no_tu1[11])) < len(type_of_atoms_table(no_tu2[11])):
        for at1 in range(0,len(type_of_atoms_table(no_tu1[11]))):
            n1=type_of_atoms_table(no_tu1[11])
            n2=type_of_atoms_table(no_tu2[11])
            numper_atom=get_perc_result(n1[at1],n2[at1])
            real_similarity.append(float(numper_atom))
    if len(type_of_atoms_table(no_tu1[11])) > len(type_of_atoms_table(no_tu2[11])):
        for at1 in range(0,len(type_of_atoms_table(no_tu2[11]))):
            n1=type_of_atoms_table(no_tu1[11])
            n2=type_of_atoms_table(no_tu2[11])
            numper_atom=get_perc_result(n1[at1],n2[at1])
            real_similarity.append(float(numper_atom))

    v = 0
    for v_v in real_similarity[0:]:
        v = int(v) + int(v_v)
    numper = len(real_similarity) * 100
    sim_res = get_perc_result(v, numper)
    print("\n")
    print("----------" + no_tu1[11] + "-------------")
    type_of_atoms_table_show(no_tu1[11])
    print("\n")
    print("----------"+no_tu2[11]+"-------------")
    type_of_atoms_table_show(no_tu2[11])
    print("\n")
    print("**","real similarity similarity percentage is :- ", int(float(sim_res)),"%")
    print("------------------------END OF REPORT--------------------------------")
    return 0






#############################
#test
def e_numper(st_name):
    tu = []
    sum = 0
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    name = str(st_name) + "_" + "Atoms"
    exe.execute('SELECT (atom_name) FROM %s  ' % (name))
    n = 0
    text = "loading ...."
    get = "...."
    for i in exe.fetchall():
        conn1 = sqlite3.connect("ASL\elemnet.db")
        exe2 = conn1.cursor()
        exe2.execute('SELECT (Z) FROM element WHERE sym = "%s" ' %(i[0:]))
        for y in exe2.fetchall():

            tu.append(float(y[0]))
    ## sum
    val = 0
    v = 0
    for u in tu[0:]:
        v = v + 1
        val += float(u)
    print("[", "e- numper calculate is  {}:".format(st_name), val, "]")


    return val
def mwt(st_name):
    tu = []
    sum = 0
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    name = str(st_name) + "_" + "Atoms"
    exe.execute('SELECT (atom_name) FROM %s  ' % (name))
    n = 0
    text = "loading ...."
    get = "...."
    for i in exe.fetchall():
        conn1 = sqlite3.connect("ASL\elemnet.db")
        exe2 = conn1.cursor()
        exe2.execute('SELECT (aw) FROM element WHERE sym = "%s" ' %(i[0:]))
        for y in exe2.fetchall():
            tu.append(float(y[0]))
    ## sum
    val = 0
    v = 0
    for u in tu[0:]:
        v = v + 1
        val += float(u)
    print("[", " mwt calculate is  {}:".format(st_name), val, "]")


    return val
def total_Density_calc(st_name):
    tu = []
    sum = 0
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    name = str(st_name) + "_" + "Atoms"
    exe.execute('SELECT (atom_name) FROM %s  ' % (name))
    n = 0
    for i in exe.fetchall():
        conn1 = sqlite3.connect("ASL\elemnet.db")
        exe2 = conn1.cursor()
        exe2.execute('SELECT (dens) FROM element WHERE sym = "%s" ' % (i[0:]))
        for y in exe2.fetchall():
            tu.append(float(y[0]))
    ## sum
    val = 0
    v = 0
    for u in tu[0:]:
        v = v + 1
        val += float(u)
    print("[", "dens calculate is  {}:".format(st_name), val, "]")

    return val
#numper of protons
def total_proton_calc(st_name):
    tu = []
    sum = 0
    conn = sqlite3.connect("ASL\st.db")
    exe = conn.cursor()
    name = str(st_name) + "_" + "Atoms"
    exe.execute('SELECT (atom_name) FROM %s  ' % (name))
    n = 0
    for i in exe.fetchall():
        conn1 = sqlite3.connect("ASL\elemnet.db")
        exe2 = conn1.cursor()
        exe2.execute('SELECT (Z) FROM element WHERE sym = "%s" ' % (i[0:]))
        for y in exe2.fetchall():
            conn3 = sqlite3.connect("ASL\elemnet.db")
            exe3 = conn3.cursor()
            exe3.execute('SELECT (aw) FROM element WHERE sym = "%s" ' % (i[0:]))
            for x in exe3.fetchall():
                tu.append(float(x[0])-float(y[0]))
    ## sum
    val = 0
    v = 0
    for u in tu[0:]:
        v = v + 1
        val += float(u)
    print("[", "proton calculate is  {}:".format(st_name), val, "]")

    return val
def change_sign(val):
    if float(val<0):
        val_final=numpy.abs(val)
        return val_final
    else:
        return val
#show tables
def show_saved_st():
    conn = sqlite3.connect("ASL\saved_names.db")
    exe = conn.cursor()
    exe.execute(""" SELECT (name) FROM saved_names""")
    n = 0
    for i in exe.fetchall():
        print(i[0:])
    return 0



    #calculation
def find_multi_numper_pattern(n1,n2,n3):
    final_sum1=0
    final_sum2 = 0
    final_sum3=0
    error_tup=[]
    #st1 n1 , n2
    st1_numper=find_pattern_numper(change_sign(n1),change_sign(n2),.00001)
    st1_error = find_pattern_error(change_sign(n1), change_sign(n2), .00001)
    error_tup.append(st1_error)
    if float(st1_error[0])>100:
        redu=float(st1_error[0])-float(100)
        redu_st=float(redu)*float(st1_numper[0])
        if float(redu_st)>float(st1_numper[0]):
            final_sum1=float(redu)-float(st1_numper[0])
        if float(redu_st) < float(st1_numper[0]):
            final_sum1 =float(st1_numper[0]) - float(redu)

    if float(st1_error[0]) < 100:
        redu = float(100) - float(st1_error[0])
        redu_st = float(redu) * float(st1_numper[0])
        if float(redu_st) > float(st1_numper[0]):
            final_sum1 = float(redu) - float(st1_numper[0])
        if float(redu_st) < float(st1_numper[0]):
            final_sum1 = float(st1_numper[0]) - float(redu)


    #st2 n2 , n3
    st1_numper2 = find_pattern_numper(change_sign(n2),change_sign(n3) , .00001)
    st1_error2 = find_pattern_error(change_sign(n2), change_sign(n3), .00001)
    if float(st1_error2[0]) > 100:
        redu2 = float(st1_error2[0]) - float(100)
        redu_st2 = float(redu2) * float(st1_numper2[0])
        if float(redu_st2) > float(st1_numper2[0]):
            final_sum2 = float(redu2) - float(st1_numper2[0])
        if float(redu_st2) < float(st1_numper2[0]):
            final_sum2 = float(st1_numper2[0]) - float(redu2)

    if float(st1_error2[0]) < 100:
        redu2 = float(100) - float(st1_error2[0])
        redu_st2 = float(redu2) * float(st1_numper2[0])
        if float(redu_st2) > float(st1_numper2[0]):
            final_sum2 = float(redu2) - float(st1_numper2[0])
        if float(redu_st2) < float(st1_numper2[0]):
            final_sum2 = float(st1_numper2[0]) - float(redu2)

    print("final :","patt 1 fraction" ,final_sum1,"patt 2 fraction",final_sum2)
    if float(final_sum1)==float(final_sum2) or float(final_sum1)-float(final_sum2)<=.9 :
        st1_numper3 = [1]
        st1_error3 = [98.0]

        if float(st1_error3[0]) > 100:
            redu3 = float(st1_error3[0]) - float(100)
            redu_st3 = float(redu3) * float(st1_numper3[0])
            print("0")
            if float(redu_st3) > float(st1_numper3[0]):
                final_sum3 = float(redu3) - float(st1_numper3[0])
                print("1")
            if float(redu_st3) < float(st1_numper3[0]):
                final_sum3 = float(st1_numper3[0]) - float(redu3)
                print("2")

        if float(st1_error3[0]) < 100:

            redu3 = float(100) - float(st1_error3[0])
            redu_st3 = float(redu3) * float(st1_numper3[0])
            if float(redu_st3) > float(st1_numper3[0]):
                final_sum3 = float(redu3) - float(st1_numper3[0])
            if float(redu_st3) < float(st1_numper3[0]):
                final_sum3 = float(st1_numper3[0]) - float(redu3)

        #
        last_factor = float(final_sum3) * float(final_sum2)
        target = float(last_factor) * float(n3)
        print("target is", change_sign(target))
        print("accuricy 1 =>", error_tup[0])
        print("accuricy 2 =>", st1_error2)
        print("accuricy 3 =>", st1_error3)
        return target


    else:
    # final_sum1 ,2
        st1_numper3 = find_pattern_numper(change_sign(float(final_sum1)), change_sign(float(final_sum2)), .00001)
        st1_error3 = find_pattern_error(change_sign(float(final_sum1)), change_sign(float(final_sum2)), .00001)

        if float(st1_error3[0]) > 100:
            redu3 = float(st1_error3[0]) - float(100)
            redu_st3 = float(redu3) * float(st1_numper3[0])
            print("0")
            if float(redu_st3) > float(st1_numper3[0]):
                final_sum3 = float(redu3) - float(st1_numper3[0])
                print("1")
            if float(redu_st3) < float(st1_numper3[0]):
                final_sum3 = float(st1_numper3[0]) - float(redu3)
                print("2")

        if float(st1_error3[0]) < 100:

            redu3 = float(100) - float(st1_error3[0])
            redu_st3 = float(redu3) * float(st1_numper3[0])
            print("10")
            if float(redu_st3) > float(st1_numper3[0]):
                final_sum3 = float(redu3) - float(st1_numper3[0])
                print("3")
            if float(redu_st3) < float(st1_numper3[0]):
                final_sum3 = float(st1_numper3[0]) - float(redu3)
                print("4")


    #
        last_factor=float(final_sum3)*float(final_sum2)
        target=float(last_factor)*float(n3)
        print("target is", change_sign(target))
        print("accuricy step 1 =>", error_tup[0])
        print("accuricy step 2 =>", st1_error2)
        print("accuricy step 3 =>", st1_error3)

        return target